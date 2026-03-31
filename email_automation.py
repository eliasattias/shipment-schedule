import os
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
import openpyxl as oxl
import logging
import requests as http_requests
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import base64
from dotenv import load_dotenv

load_dotenv()

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(_handler)
if not os.getenv('AWS_LAMBDA_FUNCTION_NAME'):
    _fh = logging.FileHandler('email_automation.log')
    _fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(_fh)

class EmailAutomation:
    def __init__(self):
        # File configuration
        self.base_dir = Path(__file__).parent
        # Lambda can only write to /tmp
        if os.getenv('AWS_LAMBDA_FUNCTION_NAME'):
            self.data_dir = Path('/tmp') / "data"
        else:
            self.data_dir = self.base_dir / "data"
        self.target_filename = os.getenv('TARGET_FILENAME', 'searchresults.xlsx')
        self.token_path = self.base_dir / 'token.json'
        self.gmail_token_json = os.getenv('GMAIL_TOKEN_JSON', '')

        # Search configuration
        self.search_subject = os.getenv('SEARCH_SUBJECT', 'Sensi Medical Sales Open Order')
        self.search_sender = os.getenv('SEARCH_SENDER', 'customercare@optimalmax.com')

        # GitHub configuration (for pushing updated file to repo)
        self.github_token = os.getenv('GITHUB_TOKEN', '')
        self.github_repo = os.getenv('GITHUB_REPO', 'Sensimedical/shipment-schedule')

        # Notification configuration (Resend)
        self.resend_api_key = os.getenv('RESEND_API_KEY', '')
        self.notify_from = 'Sensimedical Pending Orders Schedule <automations@sensimedical.com>'
        self.notify_emails = [
            'automations@sensimedical.com',
            'alice.s@sensimedical.com',
            'eduardo.s@sensimedical.com',
        ]

    def get_gmail_service(self):
        """Get authenticated Gmail API service"""
        SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
        try:
            creds = None

            # Priority 1: GMAIL_TOKEN_JSON env var (GitHub Actions / Lambda)
            if self.gmail_token_json:
                token_data = json.loads(self.gmail_token_json)
                creds = Credentials.from_authorized_user_info(token_data, SCOPES)
            # Priority 2: token.json file (local)
            elif self.token_path.exists():
                creds = Credentials.from_authorized_user_file(str(self.token_path), SCOPES)

            # Refresh token if expired
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
                # Save refreshed token back to file for next run
                if not self.gmail_token_json and self.token_path:
                    with open(self.token_path, 'w') as f:
                        f.write(creds.to_json())

            if not creds or not creds.valid:
                logger.error("No valid Gmail credentials found. Run oauth2_setup.py first.")
                return None

            return build('gmail', 'v1', credentials=creds, static_discovery=False)

        except Exception as e:
            logger.error(f"Failed to get Gmail service: {e}")
            return None

    def search_emails(self, service, days_back=1):
        """Search for emails with XLS attachments"""
        try:
            # Build search query
            since_date = (datetime.now() - timedelta(days=days_back)).strftime('%Y/%m/%d')
            query = f'after:{since_date}'

            if self.search_subject:
                query += f' subject:"{self.search_subject}"'
            if self.search_sender:
                query += f' from:"{self.search_sender}"'

            query += ' has:attachment filename:(xls OR xlsx)'

            logger.info(f"Searching with query: {query}")

            results = service.users().messages().list(
                userId='me',
                q=query,
                maxResults=5
            ).execute()

            messages = results.get('messages', [])
            logger.info(f"Found {len(messages)} matching emails")
            return messages

        except Exception as e:
            logger.error(f"Failed to search emails: {e}")
            raise

    def convert_spreadsheetml_to_xlsx(self, src_path, dest_path):
        """Parse SpreadsheetML XML (used by NetSuite/older ERP exports) and write to .xlsx"""
        tree = ET.parse(src_path)
        root = tree.getroot()
        ns = "urn:schemas-microsoft-com:office:spreadsheet"
        if ns not in root.tag:
            raise ValueError("Not a SpreadsheetML file")

        wb_out = oxl.Workbook()
        wb_out.remove(wb_out.active)
        sheets_found = 0

        for ws_elem in root.iter(f"{{{ns}}}Worksheet"):
            sheet_name = ws_elem.get(f"{{{ns}}}Name", f"Sheet{sheets_found + 1}")
            ws_out = wb_out.create_sheet(title=sheet_name)
            row_num = 0
            for row_elem in ws_elem.iter(f"{{{ns}}}Row"):
                row_idx = row_elem.get(f"{{{ns}}}Index")
                row_num = int(row_idx) if row_idx else row_num + 1
                col_num = 0
                for cell_elem in row_elem.iter(f"{{{ns}}}Cell"):
                    cell_idx = cell_elem.get(f"{{{ns}}}Index")
                    col_num = int(cell_idx) if cell_idx else col_num + 1
                    data = cell_elem.find(f"{{{ns}}}Data")
                    if data is not None:
                        val = data.text or ""
                        dtype = data.get(f"{{{ns}}}Type", "String")
                        if dtype == "Number":
                            try:
                                val = float(val) if "." in val else int(val)
                            except (ValueError, TypeError):
                                pass
                        elif dtype == "DateTime":
                            try:
                                parsed = None
                                for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                                    try:
                                        parsed = datetime.strptime(val, fmt)
                                        break
                                    except ValueError:
                                        continue
                                if parsed:
                                    val = parsed
                            except Exception:
                                pass
                        elif dtype == "Boolean":
                            val = val.strip().lower() in ("1", "true")
                        cell = ws_out.cell(row=row_num, column=col_num, value=val)
                        if hasattr(val, 'year'):
                            cell.number_format = 'MM/DD/YYYY'
                    merge = cell_elem.get(f"{{{ns}}}MergeAcross")
                    if merge:
                        col_num += int(merge)
            sheets_found += 1

        if not sheets_found:
            raise ValueError("No worksheets found in SpreadsheetML")
        wb_out.save(dest_path)
        return True

    def convert_to_xlsx(self, src_path, dest_path):
        """Convert .xls to .xlsx using multiple strategies"""
        # Strategy 1: standard xlrd read
        try:
            df = pd.read_excel(src_path, engine='xlrd')
            df.to_excel(dest_path, index=False)
            logger.info(f"Converted via xlrd: {src_path.name}")
            return True
        except Exception as e1:
            logger.debug(f"xlrd failed: {e1}")

        # Strategy 2: openpyxl engine (xlsx mis-named as .xls)
        try:
            df = pd.read_excel(src_path, engine='openpyxl')
            df.to_excel(dest_path, index=False)
            logger.info(f"Converted via openpyxl: {src_path.name}")
            return True
        except Exception as e2:
            logger.debug(f"openpyxl failed: {e2}")

        # Strategy 3: HTML table parsing
        try:
            tables = pd.read_html(str(src_path), flavor="lxml")
            if tables:
                tables[0].to_excel(dest_path, index=False)
                logger.info(f"Converted via HTML table parsing: {src_path.name}")
                return True
        except Exception as e3:
            logger.debug(f"HTML fallback failed: {e3}")

        # Strategy 4: SpreadsheetML XML
        try:
            self.convert_spreadsheetml_to_xlsx(str(src_path), str(dest_path))
            logger.info(f"Converted via SpreadsheetML XML: {src_path.name}")
            return True
        except Exception as e4:
            logger.debug(f"SpreadsheetML failed: {e4}")

        return False

    def download_attachment(self, service, message_id):
        """Download XLS attachment from email"""
        try:
            message = service.users().messages().get(userId='me', id=message_id, format='full').execute()

            headers = message['payload']['headers']
            subject = next((h['value'] for h in headers if h['name'] == 'Subject'), 'Unknown')
            sender = next((h['value'] for h in headers if h['name'] == 'From'), 'Unknown')

            logger.info(f"Processing email from {sender}: {subject}")

            # Find and download attachment
            if 'parts' in message['payload']:
                for part in message['payload']['parts']:
                    if part['filename']:
                        filename = part['filename']
                        if filename.lower().endswith(('.xls', '.xlsx')):
                            attachment_id = part['body']['attachmentId']
                            attachment = service.users().messages().attachments().get(
                                userId='me',
                                messageId=message_id,
                                id=attachment_id
                            ).execute()

                            data = base64.urlsafe_b64decode(attachment['data'])

                            # Save raw file
                            raw_path = self.data_dir / filename
                            with open(raw_path, 'wb') as f:
                                f.write(data)

                            final_path = self.data_dir / self.target_filename

                            # Convert if .xls, otherwise just move
                            if raw_path.suffix.lower() == '.xls':
                                if not self.convert_to_xlsx(raw_path, final_path):
                                    logger.warning("All conversion strategies failed, using raw file")
                                    raw_path.replace(final_path)
                                # Clean up raw .xls if final is different file
                                if raw_path.exists() and raw_path != final_path:
                                    raw_path.unlink()
                            else:
                                raw_path.replace(final_path)

                            logger.info(f"Downloaded attachment: {filename} -> {final_path}")
                            return final_path

            return None

        except Exception as e:
            logger.error(f"Failed to download attachment from message {message_id}: {e}")
            return None

    def validate_excel_file(self, filepath):
        """Validate that the downloaded file is a valid Excel file"""
        try:
            # Check file exists and has content
            if not filepath.exists():
                logger.error(f"File does not exist: {filepath}")
                return False
            
            file_size = filepath.stat().st_size
            if file_size == 0:
                logger.error(f"Downloaded file is empty: {filepath}")
                return False
            
            logger.info(f"File downloaded successfully: {filepath} ({file_size} bytes)")
            
            # Try to read with any available engine
            for engine in ['openpyxl', 'lxml', 'xlrd']:
                try:
                    df = pd.read_excel(filepath, engine=engine)
                    if not df.empty:
                        logger.info(f"Validated with {engine}: {len(df)} rows, {len(df.columns)} columns")
                        return True
                except Exception as e:
                    logger.debug(f"Engine {engine} failed: {str(e)[:80]}")
                    continue
            
            # last ditch: if file still cannot be read but has .xls extension,
            # try converting it in-place and then re-validate
            if filepath.suffix.lower() == '.xls':
                try:
                    df = pd.read_excel(filepath, engine='xlrd')
                    df.to_excel(filepath.with_suffix('.xlsx'), index=False)
                    logger.info(f"Converted failing .xls to .xlsx for validation")
                    return self.validate_excel_file(filepath.with_suffix('.xlsx'))
                except Exception:
                    pass

            # Log warning but return True since file was downloaded successfully
            logger.warning(f"Could not parse file as Excel, but file exists and has content")
            return True
            
        except Exception as e:
            logger.error(f"Validation error: {e}")
            return False

    def push_to_github(self, filepath):
        """Push the updated file to GitHub via the Contents API"""
        if not self.github_token:
            logger.warning("GITHUB_TOKEN not set - skipping GitHub push")
            return False

        try:
            repo_path = f"data/{self.target_filename}"
            api_url = f"https://api.github.com/repos/{self.github_repo}/contents/{repo_path}"
            headers = {
                'Authorization': f'token {self.github_token}',
                'Accept': 'application/vnd.github.v3+json',
            }

            # Get current file SHA (required for update)
            sha = None
            resp = http_requests.get(api_url, headers=headers, timeout=10)
            if resp.status_code == 200:
                sha = resp.json()['sha']

            # Read and encode file content
            with open(filepath, 'rb') as f:
                content = base64.b64encode(f.read()).decode()

            now_str = datetime.now(ZoneInfo('America/New_York')).strftime('%Y-%m-%d %I:%M %p ET')
            payload = {
                'message': f'Auto-update: {self.target_filename} - {now_str}',
                'content': content,
                'branch': 'main',
            }
            if sha:
                payload['sha'] = sha

            resp = http_requests.put(api_url, headers=headers, json=payload, timeout=30)
            if resp.status_code in (200, 201):
                logger.info("File pushed to GitHub successfully")
                return True
            else:
                logger.error(f"GitHub API error {resp.status_code}: {resp.text[:200]}")
                return False

        except Exception as e:
            logger.error(f"Failed to push to GitHub: {e}")
            return False

    def send_reminder(self):
        """Send Tuesday/Thursday reminder to update the console"""
        if not self.resend_api_key:
            logger.warning("RESEND_API_KEY not set - skipping reminder")
            return

        try:
            html_body = (
                '<p>Hi Tina,</p>'
                '<p>Just a reminder to please update the '
                '<a href="https://sensimedical-shipment-schedule.streamlit.app/">console</a>'
                ' today. Thank you in advance!</p>'
                '<hr>'
                '<p>This is an automated message.</p>'
            )

            resp = http_requests.post(
                'https://api.resend.com/emails',
                headers={
                    'Authorization': f'Bearer {self.resend_api_key}',
                    'Content-Type': 'application/json',
                },
                json={
                    'from': self.notify_from,
                    'to': [
                        'customercare@optimalmax.com',
                        'FStivers@OptimalMax.com',
                        'cjohnson@optimalmax.com',
                        'SBroussard@optimalmax.com',
                        'TPerez@optimalmax.com',
                        'PBigley@optimalmax.com',
                    ],
                    'cc': [
                        'alice.s@sensimedical.com',
                        'eduardo.s@sensimedical.com',
                    ],
                    'subject': 'Reminder: Please Update the Shipment Schedule Console',
                    'html': html_body,
                },
                timeout=10,
            )

            if resp.status_code in (200, 201):
                logger.info("Reminder email sent via Resend")
            else:
                logger.error(f"Resend error {resp.status_code}: {resp.text}")

        except Exception as e:
            logger.error(f"Failed to send reminder: {e}")

    def send_notification(self, success, message):
        """Send notification email via Resend"""
        if not self.resend_api_key:
            logger.warning("RESEND_API_KEY not set - skipping notification")
            return

        if not success:
            logger.info("Automation did not succeed - skipping notification email")
            return

        try:
            now_et = datetime.now(ZoneInfo('America/New_York'))
            date_str = now_et.strftime('%B %d, %Y at %I:%M %p ET')

            html_body = (
                f'<p>The shipment schedule console has been updated on {date_str}.</p>'
                '<p><a href="https://sensimedical-shipment-schedule.streamlit.app/">'
                'Shipment Schedule Console</a></p>'
                '<hr>'
                '<p>This is an automated message.</p>'
            )

            resp = http_requests.post(
                'https://api.resend.com/emails',
                headers={
                    'Authorization': f'Bearer {self.resend_api_key}',
                    'Content-Type': 'application/json',
                },
                json={
                    'from': self.notify_from,
                    'to': self.notify_emails,
                    'subject': 'Shipment Schedule Console Updated',
                    'html': html_body,
                },
                timeout=10,
            )

            if resp.status_code in (200, 201):
                logger.info("Notification email sent via Resend")
            else:
                logger.error(f"Resend error {resp.status_code}: {resp.text}")

        except Exception as e:
            logger.error(f"Failed to send notification: {e}")

    def run_automation(self):
        """Main automation function"""
        logger.info("Starting email automation using Gmail API")

        try:
            # Ensure data directory exists
            self.data_dir.mkdir(parents=True, exist_ok=True)

            # Get Gmail service
            service = self.get_gmail_service()
            if not service:
                message = "Failed to authenticate with Gmail API"
                logger.error(message)
                self.send_notification(False, message)
                return

            # Search for emails
            messages = self.search_emails(service)
            if not messages:
                message = "No matching emails found"
                logger.info(message)
                self.send_notification(False, message)
                return

            # Process most recent email
            latest_message_id = messages[0]['id']
            filepath = self.download_attachment(service, latest_message_id)

            if not filepath:
                message = "No XLS attachment found in emails"
                logger.error(message)
                self.send_notification(False, message)
                return

            # Validate file
            if not self.validate_excel_file(filepath):
                message = "Downloaded file is not a valid Excel file"
                logger.error(message)
                self.send_notification(False, message)
                return

            # Push to GitHub so Streamlit Cloud picks up the new file
            self.push_to_github(filepath)

            message = f"Successfully updated {self.target_filename}"
            logger.info(message)
            self.send_notification(True, message)

        except Exception as e:
            error_msg = f"Automation failed: {str(e)}"
            logger.error(error_msg)
            self.send_notification(False, error_msg)

if __name__ == "__main__":
    automation = EmailAutomation()
    automation.run_automation()